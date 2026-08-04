"""ATIS Gold Desk — FastAPI backend connected to every engine."""

from __future__ import annotations

import asyncio
import json
import os
from contextlib import asynccontextmanager
from pathlib import Path
from typing import Any

import pandas as pd
import yaml
from fastapi import FastAPI, HTTPException
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

from atis.config import (
    CONFIG_DIR,
    PROJECT_ROOT,
    clear_config_caches,
    ensure_project_dirs,
    get_path,
    load_engine_config,
    load_symbols,
    load_timeframes,
    read_secrets_env,
    save_mt5_credentials,
)
from atis.engines.engine1_ingestion import run_ingestion
from atis.engines.engine2_cleaning import run_cleaning
from atis.engines.engine3_features import run_features
from atis.engines.engine4_training import run_training
from atis.engines.engine4_training.data_sources import training_source_meta
from atis.engines.engine5_live_trading import (
    close_position,
    close_positions_filtered,
    list_open_positions,
    run_live_once,
)
from atis.shared.data_registry import DataStateRegistry
from atis.shared.data_json import load_timeframe_json
from atis.shared.feature_engine.patterns import (
    bullish_keys,
    bearish_keys,
    pattern_category_map,
    pattern_labels,
)
from atis.shared.mt5_client import MT5Client, mt5_session, ping_mt5
from atis.shared.pattern_discovery import run_pattern_discovery, export_pattern_json_from_kb
from atis.shared.pattern_kb import PatternKnowledgeBase
from atis.shared.pattern_store import (
    SECTIONS,
    list_pattern_files,
    load_section,
    patterns_root,
    rebuild_relations_from_features,
    relations_has_content,
    save_relations_section,
)
from atis.web.autotrader import autotrader
from atis.web.jobs import jobs
from atis.web.position_watcher import position_watcher

ensure_project_dirs()

STATIC_DIR = Path(__file__).resolve().parent / "static"

# Kept for process lifetime so autotrade/engine cycles share one IPC link
_mt5_keepalive: MT5Client | None = None


@asynccontextmanager
async def _lifespan(_app: FastAPI):
    global _mt5_keepalive
    client = MT5Client()
    try:
        client.connect()
        _mt5_keepalive = client
    except Exception:
        _mt5_keepalive = None
    position_watcher.set_symbol_provider(lambda: _primary()[0])
    position_watcher.start()
    try:
        yield
    finally:
        position_watcher.stop()
        autotrader.stop()
        if _mt5_keepalive is not None:
            try:
                _mt5_keepalive.disconnect()
            except Exception:
                pass
            _mt5_keepalive = None


def _mt5_status(*, auto_reconnect: bool = False) -> dict[str, Any]:
    """Return MT5 status and optionally heal the shared connection first."""
    global _mt5_keepalive

    def _payload_from(client: MT5Client) -> dict[str, Any]:
        account = client.account_summary()
        terminal = client.terminal_info()
        return {
            "ok": True,
            "login": account.get("login"),
            "server": account.get("server"),
            "balance": account.get("balance"),
            "currency": account.get("currency"),
            "trade_allowed": account.get("trade_allowed"),
            "terminal_connected": terminal.get("connected"),
            "terminal_name": terminal.get("name"),
        }

    reconnected = False
    try:
        if _mt5_keepalive is not None and _mt5_keepalive.connected:
            payload = _payload_from(_mt5_keepalive)
            payload["reconnected"] = False
            return payload

        if auto_reconnect:
            if _mt5_keepalive is None:
                _mt5_keepalive = MT5Client()
            _mt5_keepalive.connect()
            payload = _payload_from(_mt5_keepalive)
            payload["reconnected"] = True
            return payload

        payload = ping_mt5()
        payload["reconnected"] = False
        return payload
    except Exception as exc:
        if auto_reconnect and not reconnected:
            try:
                _mt5_keepalive = MT5Client()
                _mt5_keepalive.connect()
                payload = _payload_from(_mt5_keepalive)
                payload["reconnected"] = True
                return payload
            except Exception as retry_exc:
                return {"ok": False, "error": str(retry_exc), "reconnected": False}
        return {"ok": False, "error": str(exc), "reconnected": False}

PATTERN_LABELS = pattern_labels()
PATTERN_CATEGORY = pattern_category_map()
BULLISH_PATS = bullish_keys()
BEARISH_PATS = bearish_keys()

app = FastAPI(title="ATIS Gold Desk", version="0.3.0", lifespan=_lifespan)
app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")


def _trading() -> dict[str, Any]:
    return load_engine_config().get("trading", {})


def _primary() -> tuple[str, str]:
    t = _trading()
    return (
        str(t.get("primary_symbol", "XAUUSD")),
        str(t.get("primary_timeframe", "H1")),
    )


_primary_symbol, _ = _primary()
DataStateRegistry().prune_symbols([_primary_symbol])


def _read_json(path: Path) -> Any:
    if not path.exists():
        return None
    return json.loads(path.read_text(encoding="utf-8"))


def _reconcile_regime_validation(regime: dict[str, Any] | None) -> dict[str, Any]:
    """Fix legacy false-unstable labels when all regimes stay strongly profitable."""
    if not isinstance(regime, dict) or not regime:
        return regime or {}
    try:
        from atis.engines.engine4_training.validation_protocols import reconcile_regime_stability

        return reconcile_regime_stability(regime)
    except Exception:
        return regime


def _reconcile_tf_row_regime(row: dict[str, Any] | None) -> dict[str, Any]:
    """Apply regime reconcile on a timeframe status/matrix row (live job or report)."""
    if not isinstance(row, dict) or not row:
        return row or {}
    out = dict(row)
    nested = out.get("metrics") if isinstance(out.get("metrics"), dict) else None
    if nested and isinstance(nested.get("regime_validation"), dict):
        metrics = dict(nested)
        metrics["regime_validation"] = _reconcile_regime_validation(metrics.get("regime_validation"))
        out["metrics"] = metrics
    if isinstance(out.get("regime_validation"), dict):
        out["regime_validation"] = _reconcile_regime_validation(out.get("regime_validation"))
    return out


def _reconcile_job_details(details: dict[str, Any] | None) -> dict[str, Any]:
    """Live training cards read job.details.timeframes — reconcile before UI sees them."""
    if not isinstance(details, dict) or not details:
        return details or {}
    out = dict(details)
    tfs = out.get("timeframes")
    if isinstance(tfs, dict) and tfs:
        out["timeframes"] = {
            str(k): _reconcile_tf_row_regime(v if isinstance(v, dict) else {})
            for k, v in tfs.items()
        }
    return out


def _read_jsonl_tail(path: Path, limit: int = 1) -> list[Any]:
    if not path.exists():
        return []
    lines = [line for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]
    rows = []
    for line in lines[-limit:]:
        try:
            rows.append(json.loads(line))
        except json.JSONDecodeError:
            continue
    return rows


def _llmodel_meta() -> dict[str, Any] | None:
    models_dir = get_path("models")
    artifact = models_dir / "LLModel"
    meta = models_dir / "LLModel.meta.json"
    metrics = models_dir / "LLModel.metrics.json"
    if not artifact.exists():
        return None
    return {
        "artifact_path": str(artifact),
        "metadata": _read_json(meta),
        "metrics": _read_json(metrics),
        "exists": True,
    }


def _final_model_meta() -> dict[str, Any] | None:
    models_dir = get_path("models")
    meta = _read_json(models_dir / "FinalModel.meta.json") or _read_json(models_dir / "FinalModel" / "FINAL_MODEL.json")
    if not meta or not meta.get("exists"):
        artifact = models_dir / "FinalModel" / "model.joblib"
        if not artifact.exists():
            return None
        return {
            "exists": True,
            "artifact_path": str(artifact),
            "artifact_dir": str(models_dir / "FinalModel"),
            "mode": "paper_only",
        }
    return meta


def _latest_model_dir(symbol: str, timeframe: str) -> Path | None:
    base = get_path("models") / symbol / timeframe
    champ = base / "champion.json"
    if champ.exists():
        meta = _read_json(champ) or {}
        mp = meta.get("model_path")
        if mp:
            p = Path(mp).parent
            if p.exists():
                return p
    if not base.exists():
        return None
    versions = sorted([p for p in base.iterdir() if p.is_dir()], reverse=True)
    return versions[0] if versions else None


def _parquet_meta(path: Path) -> dict[str, Any] | None:
    if not path.exists():
        return None
    try:
        df = pd.read_parquet(path, columns=["timestamp"])
        if df.empty:
            return {"exists": True, "rows": 0}
        ts = pd.to_datetime(df["timestamp"], utc=True)
        return {
            "exists": True,
            "rows": int(len(df)),
            "first_ts": str(ts.min()),
            "last_ts": str(ts.max()),
            "days": float((ts.max() - ts.min()).total_seconds() / 86400.0),
            "path": str(path),
        }
    except Exception as exc:
        return {"exists": True, "error": str(exc), "path": str(path)}


class RunRequest(BaseModel):
    symbols: list[str] | None = None
    timeframes: list[str] | None = None
    force_rebuild: bool = False
    resume: bool = True


class LiveRequest(BaseModel):
    symbols: list[str] | None = None
    timeframe: str | None = None
    execute_demo: bool = False
    allow_ungated: bool = True


class KillSwitchRequest(BaseModel):
    active: bool = True
    reason: str = "manual"


class ClosePositionsRequest(BaseModel):
    """Close one ticket, or filter: all | winners | losers."""

    ticket: int | None = None
    mode: str | None = None  # all | winners | losers
    symbol: str | None = None
    atis_only: bool = True


class AutoCloseSettingsRequest(BaseModel):
    """Toggle automatic profit/loss exits on open positions."""

    enabled: bool | None = None
    min_profit: float | None = None
    loss_enabled: bool | None = None
    max_loss: float | None = None


class DeleteRlEpisodesRequest(BaseModel):
    episode_ids: list[str] = []


class AutoTradeRequest(BaseModel):
    mode: str = "paper"  # paper | demo
    interval_seconds: int = 60
    symbol: str | None = None
    timeframe: str | None = None
    timeframes: list[str] | None = None
    # When multiple TFs are selected, default is one decision/order per TF.
    multi_tf_independent: bool | None = True
    fusion_mode: str | None = None  # independent | weighted_consensus | majority | ...


class LiveSettingsRequest(BaseModel):
    use_live_spread_filter: bool | None = None
    max_entry_spread_pips: float | None = None
    tight_spread_pips: float | None = None
    max_entries_per_cycle: int | None = None


class MT5CredentialsRequest(BaseModel):
    login: int | str | None = None
    password: str | None = None
    server: str | None = None
    path: str | None = None
    reconnect: bool = True


class TrainingSettingsRequest(BaseModel):
    settings: dict[str, Any]


class OpenPathRequest(BaseModel):
    path: str


def _live_settings_payload(cfg: dict[str, Any] | None = None) -> dict[str, Any]:
    live = cfg if cfg is not None else (load_engine_config().get("engine5_live") or {})
    return {
        "use_live_spread_filter": bool(live.get("use_live_spread_filter", True)),
        "max_entry_spread_pips": float(live.get("max_entry_spread_pips", 12.0)),
        "tight_spread_pips": float(live.get("tight_spread_pips", 12.0)),
        "max_entries_per_cycle": int(live.get("max_entries_per_cycle", 8)),
        "max_open_positions": int(live.get("max_open_positions", 20)),
        "mode_label_ar": (
            "موديل ثم سبريد"
            if bool(live.get("use_live_spread_filter", True))
            else "موديل فقط"
        ),
    }


def _training_settings_payload(cfg: dict[str, Any] | None = None) -> dict[str, Any]:
    train = cfg if cfg is not None else (load_engine_config().get("engine4_training") or {})
    if not isinstance(train, dict):
        train = {}
    return {
        "settings": train,
        "count": len(train),
        "section": "engine4_training",
    }


def _mt5_settings_payload() -> dict[str, Any]:
    secrets = read_secrets_env()
    login = secrets.get("MT5_LOGIN") or os.environ.get("MT5_LOGIN") or ""
    server = secrets.get("MT5_SERVER") or os.environ.get("MT5_SERVER") or ""
    path = secrets.get("MT5_PATH") or os.environ.get("MT5_PATH") or ""
    password = secrets.get("MT5_PASSWORD") or os.environ.get("MT5_PASSWORD") or ""
    configured = bool(login and password and server)
    return {
        "login": login,
        "server": server,
        "path": path,
        "password_set": bool(password),
        "password_masked": "********" if password else "",
        "configured": configured,
        "secrets_file": str(CONFIG_DIR / "secrets.env"),
    }


def _reconnect_mt5_keepalive() -> dict[str, Any]:
    """Drop shared MT5 link and reconnect with freshly loaded credentials."""
    global _mt5_keepalive
    if _mt5_keepalive is not None:
        try:
            _mt5_keepalive.disconnect()
        except Exception:
            pass
        _mt5_keepalive = None
    client = MT5Client()
    client.connect()
    _mt5_keepalive = client
    account = client.account_summary()
    return {
        "ok": True,
        "login": account.get("login"),
        "server": account.get("server"),
        "balance": account.get("balance"),
        "currency": account.get("currency"),
        "trade_allowed": account.get("trade_allowed"),
    }


@app.get("/")
def index() -> FileResponse:
    return FileResponse(
        STATIC_DIR / "index.html",
        headers={
            "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
            "Pragma": "no-cache",
        },
    )


@app.get("/api/health")
def health() -> dict[str, Any]:
    symbol, tf = _primary()
    cfg = load_engine_config()
    return {
        "ok": True,
        "project": cfg.get("project"),
        "trading": cfg.get("trading"),
        "primary_symbol": symbol,
        "primary_timeframe": tf,
        "instrument": cfg.get("trading", {}).get("instrument_name", "Gold"),
    }


@app.post("/api/system/open-path")
def system_open_path(req: OpenPathRequest) -> dict[str, Any]:
    path = Path(req.path).expanduser()
    if not path.exists():
        raise HTTPException(404, f"Path not found: {path}")
    try:
        os.startfile(str(path))  # type: ignore[attr-defined]
    except AttributeError as exc:
        raise HTTPException(400, "فتح المسار مدعوم حاليًا على Windows فقط") from exc
    except OSError as exc:
        raise HTTPException(500, f"تعذر فتح المسار: {exc}") from exc
    return {"ok": True, "path": str(path)}


@app.get("/api/config")
def api_config() -> dict[str, Any]:
    cfg = load_engine_config()
    return {
        "engine_config": cfg,
        "symbols": load_symbols(),
        "timeframes": list(load_timeframes().keys()),
    }


@app.get("/api/mt5/status")
def mt5_status() -> dict[str, Any]:
    return _mt5_status(auto_reconnect=True)


@app.get("/api/registry")
def registry(layer: str | None = None) -> dict[str, Any]:
    reg = DataStateRegistry()
    if layer:
        rows = [r.__dict__ for r in reg.list_layer(layer)]
    else:
        rows = []
        for lyr in ("raw", "clean", "features"):
            rows.extend(r.__dict__ for r in reg.list_layer(lyr))
    symbol, _ = _primary()
    gold = [r for r in rows if r.get("symbol") == symbol]
    return {
        "rows": rows,
        "gold": gold,
        "count": len(rows),
        "root": str(reg.root_dir),
        "files": reg.list_state_files(),
    }


@app.get("/api/registry/files")
def registry_files() -> dict[str, Any]:
    """List per-timeframe data-state JSON files."""
    reg = DataStateRegistry()
    files = reg.list_state_files()
    return {
        "root": str(reg.root_dir),
        "files": files,
        "count": len(files),
    }


@app.get("/api/registry/files/{timeframe}")
def registry_file_content(timeframe: str) -> dict[str, Any]:
    """Return one timeframe state JSON document for in-UI viewing."""
    reg = DataStateRegistry()
    path = reg.timeframe_path(timeframe)
    doc = reg.load_timeframe_doc(timeframe)
    if doc is None:
        raise HTTPException(404, f"No state JSON for timeframe {timeframe}")
    return {
        "timeframe": timeframe,
        "filename": path.name,
        "path": str(path),
        "relative_path": f"data/registry/{path.name}",
        "content": doc,
    }


@app.get("/api/data/coverage")
def data_coverage() -> dict[str, Any]:
    """Full gold data status across all timeframes and layers."""
    symbol, primary_tf = _primary()
    tfs = list(load_timeframes().keys())
    coverage: list[dict[str, Any]] = []
    reg = DataStateRegistry()

    for tf in tfs:
        state_path = reg.timeframe_path(tf)
        item: dict[str, Any] = {
            "symbol": symbol,
            "timeframe": tf,
            "layers": {},
            "state_json": {
                "exists": state_path.exists(),
                "filename": state_path.name,
                "path": str(state_path),
                "relative_path": f"data/registry/{state_path.name}",
                "size_bytes": int(state_path.stat().st_size) if state_path.exists() else 0,
            },
        }
        paths = {
            "raw": get_path("data_raw") / symbol / tf / f"{symbol}_{tf}.parquet",
            "clean": get_path("data_clean") / symbol / tf / f"{symbol}_{tf}.parquet",
            "features": get_path("data_features") / symbol / tf / "features.parquet",
        }
        for layer, path in paths.items():
            meta = _parquet_meta(path)
            reg_row = reg.get(symbol, tf, layer)
            item["layers"][layer] = {
                "file": meta,
                "registry": reg_row.__dict__ if reg_row else None,
            }
        coverage.append(item)

    return {
        "symbol": symbol,
        "primary_timeframe": primary_tf,
        "timeframes": tfs,
        "coverage": coverage,
        "state_files": reg.list_state_files(),
        "registry_root": str(reg.root_dir),
    }


@app.get("/api/patterns")
def patterns(limit: int = 500, lookback: int = 5000, timeframe: str | None = None) -> dict[str, Any]:
    """Advanced candlestick + chart + compound pattern detections for gold."""
    symbol, primary_tf = _primary()
    tf = timeframe or primary_tf
    path = get_path("data_features") / symbol / tf / "features.parquet"
    json_path = get_path("data_features") / symbol / tf / "features.json"
    json_files = list_pattern_files(symbol=symbol)
    json_files_tf = [f for f in json_files if f.get("timeframe") == tf]
    empty_payload = {
        "symbol": symbol,
        "timeframe": tf,
        "detections": [],
        "counts": {},
        "candle_counts": {},
        "chart_counts": {},
        "compound_counts": {},
        "active_now": [],
        "structure": {},
        "bias_label": "محايد",
        "catalog_size": len(pattern_labels()),
        "patterns_with_hits": 0,
        "total_detections": 0,
        "knowledge": PatternKnowledgeBase().summary(symbol, tf),
        "json_files": json_files,
        "json_files_current_tf": json_files_tf,
        "patterns_root": str(patterns_root()),
        "ohlc": [],
        "markers": [],
        "empty": True,
        "relations": {
            "summary": "لا علاقات بعد — أعد بناء الشبكة",
            "edges": [],
            "nodes": [],
            "sequences": [],
            "counts": {},
            "empty": True,
        },
    }
    if not json_path.exists() and not path.exists():
        return empty_payload

    # Prefer parquet for the UI — JSON can be hundreds of MB and blocks other APIs.
    if path.exists():
        df = pd.read_parquet(path)
    else:
        df = load_timeframe_json(json_path)
    if lookback and lookback > 0:
        df = df.tail(lookback).copy()
    else:
        df = df.copy()
    df["timestamp"] = pd.to_datetime(df["timestamp"], utc=True)
    labels = pattern_labels()
    cats = pattern_category_map()
    bull = bullish_keys()
    bear = bearish_keys()
    pat_cols = [
        c
        for c in df.columns
        if c in labels
        or (
            (c.startswith("pat_") or c.startswith("cmp_") or c.startswith("disc_"))
            and c not in {"pat_bias", "pat_strength"}
        )
    ]

    detections: list[dict[str, Any]] = []
    counts: dict[str, int] = {}
    candle_counts: dict[str, int] = {}
    chart_counts: dict[str, int] = {}
    compound_counts: dict[str, int] = {}

    kb_stats = {
        r["pattern_key"]: r
        for r in PatternKnowledgeBase().list_stats(symbol, tf, min_occurrences=1, limit=2000)
    }

    for col in pat_cols:
        hits = df[df[col] == 1]
        label = labels.get(col, col)
        counts[label] = int(len(hits))
        cat = cats.get(col, "candle")
        if col.startswith("disc_") or col.startswith("cmp_") or cat == "compound":
            compound_counts[label] = int(len(hits))
            cat = "compound"
        elif cat == "chart":
            chart_counts[label] = int(len(hits))
        else:
            candle_counts[label] = int(len(hits))
        # Keep many recent hits per pattern (was 8) — diversity over truncation
        for _, row in hits.tail(40).iterrows():
            bias = "bullish" if col in bull else ("bearish" if col in bear else "neutral")
            st = kb_stats.get(col) or {}
            detections.append(
                {
                    "timestamp": str(row["timestamp"]),
                    "pattern": label,
                    "pattern_key": col,
                    "category": cat,
                    "bias": bias,
                    "close": float(row["close"]),
                    "open": float(row["open"]),
                    "high": float(row["high"]),
                    "low": float(row["low"]),
                    "strength": float(row["pat_strength"])
                    if "pat_strength" in row and pd.notna(row.get("pat_strength"))
                    else None,
                    "occurrences": st.get("occurrences"),
                    "success_rate": st.get("success_rate"),
                    "confidence": st.get("confidence"),
                    "conditions": st.get("conditions") or st.get("catalog_conditions"),
                }
            )

    detections.sort(key=lambda x: x["timestamp"], reverse=True)
    detections = detections[: max(limit, 500)]

    last = df.iloc[-1]
    active_now = [
        {
            "pattern": labels.get(c, c),
            "key": c,
            "category": cats.get(c, "candle"),
            "bias": "bullish" if c in bull else ("bearish" if c in bear else "neutral"),
        }
        for c in pat_cols
        if float(last[c]) == 1.0
    ]

    structure = {
        "support_level": float(last["support_level"]) if "support_level" in df.columns and pd.notna(last.get("support_level")) else None,
        "resist_level": float(last["resist_level"]) if "resist_level" in df.columns and pd.notna(last.get("resist_level")) else None,
        "dist_to_support": float(last["dist_to_support"]) if "dist_to_support" in df.columns and pd.notna(last.get("dist_to_support")) else None,
        "dist_to_resist": float(last["dist_to_resist"]) if "dist_to_resist" in df.columns and pd.notna(last.get("dist_to_resist")) else None,
        "session": str(last["session"]) if "session" in df.columns else None,
        "vol_regime": str(last["vol_regime"]) if "vol_regime" in df.columns else None,
        "trend_strength": float(last["trend_strength"]) if "trend_strength" in df.columns and pd.notna(last.get("trend_strength")) else None,
        "rsi_14": float(last["rsi_14"]) if "rsi_14" in df.columns and pd.notna(last.get("rsi_14")) else None,
        "atr": float(last["atr"]) if "atr" in df.columns and pd.notna(last.get("atr")) else None,
        "pat_bias": int(last["pat_bias"]) if "pat_bias" in df.columns and pd.notna(last.get("pat_bias")) else 0,
        "pat_strength": float(last["pat_strength"]) if "pat_strength" in df.columns and pd.notna(last.get("pat_strength")) else None,
        "chart_pattern_score": float(last["chart_pattern_score"]) if "chart_pattern_score" in df.columns and pd.notna(last.get("chart_pattern_score")) else 0.0,
        "structure_hh_hl": float(last["structure_hh_hl"]) if "structure_hh_hl" in df.columns and pd.notna(last.get("structure_hh_hl")) else 0.0,
        "trendline_slope": float(last["trendline_slope"]) if "trendline_slope" in df.columns and pd.notna(last.get("trendline_slope")) else None,
    }

    chart_n = min(120, len(df))
    chart_df = df.tail(chart_n)
    markers = []
    for _, row in chart_df.iterrows():
        hits = [labels.get(c, c) for c in pat_cols if float(row[c]) == 1.0]
        if hits:
            markers.append(
                {
                    "timestamp": str(row["timestamp"]),
                    "price": float(row["high"]),
                    "patterns": hits[:8],
                    "bias": int(row["pat_bias"]) if "pat_bias" in row and pd.notna(row.get("pat_bias")) else 0,
                }
            )

    ohlc = chart_df[["timestamp", "open", "high", "low", "close"]].copy()
    ohlc["timestamp"] = ohlc["timestamp"].astype(str)

    bias_label = "محايد"
    if structure["pat_bias"] > 0 or structure["chart_pattern_score"] > 0:
        bias_label = "صاعد"
    elif structure["pat_bias"] < 0 or structure["chart_pattern_score"] < 0:
        bias_label = "هابط"

    kb = PatternKnowledgeBase().summary(symbol, tf)

    def _sec(name: str) -> dict[str, Any]:
        return load_section(symbol, tf, name) or {}

    new_sec = _sec("new_patterns")
    rank_sec = _sec("rankings")
    rel_sec = _sec("relations")
    val_sec = _sec("validation_report")

    # Lazy rebuild when relations.json is an empty shell (common after resume bug)
    if not relations_has_content(rel_sec):
        try:
            from atis.shared.pattern_discovery.relations import (
                build_pattern_relations,
                pattern_relation_columns,
            )

            rebuilt = build_pattern_relations(
                df, pattern_relation_columns(df), labels=labels
            )
            if relations_has_content(rebuilt):
                save_relations_section(
                    symbol=symbol,
                    timeframe=tf,
                    relations=rebuilt,
                    bars_scanned=int(len(df)),
                )
                rel_sec = {**rebuilt, "rebuilt": True}
            else:
                rel_sec = rebuilt or rel_sec
        except Exception:
            pass

    counts_rel = rel_sec.get("counts") or {}
    if not counts_rel and rel_sec.get("edges"):
        edges_all = rel_sec.get("edges") or []
        counts_rel = {
            "co_occurrence": sum(1 for e in edges_all if e.get("relation") == "co_occurrence"),
            "precedes": sum(1 for e in edges_all if e.get("relation") == "precedes"),
            "cancels": sum(1 for e in edges_all if e.get("relation") == "cancels"),
        }

    return {
        "symbol": symbol,
        "timeframe": tf,
        "lookback_bars": int(len(df)),
        "counts": counts,
        "candle_counts": candle_counts,
        "chart_counts": chart_counts,
        "compound_counts": compound_counts,
        "detections": detections,
        "active_now": active_now,
        "structure": structure,
        "bias_label": bias_label,
        "catalog_size": len(labels),
        "patterns_with_hits": len([v for v in counts.values() if v > 0]),
        "total_detections": int(sum(counts.values())),
        "knowledge": kb,
        "json_files": json_files,
        "json_files_current_tf": json_files_tf,
        "patterns_root": str(patterns_root()),
        "ohlc": ohlc.to_dict(orient="records"),
        "markers": markers,
        "empty": False,
        "new_patterns": {
            "count": new_sec.get("count") or len(new_sec.get("items") or []),
            "approved": new_sec.get("approved"),
            "rejected": new_sec.get("rejected"),
            "items": (new_sec.get("items") or [])[:20],
        },
        "rankings": {
            "engine4_recommended": (rank_sec.get("engine4_recommended") or [])[:15],
        },
        "relations": {
            "summary": rel_sec.get("summary") or "لا علاقات بعد — أعد بناء الشبكة",
            "edges": (rel_sec.get("edges") or [])[:60],
            "nodes": (rel_sec.get("nodes") or [])[:40],
            "sequences": (rel_sec.get("sequences") or [])[:20],
            "counts": counts_rel,
            "bars": rel_sec.get("bars") or rel_sec.get("bars_scanned"),
            "lag_max": rel_sec.get("lag_max"),
            "active_patterns": rel_sec.get("active_patterns"),
            "rebuilt": bool(rel_sec.get("rebuilt")),
            "empty": not relations_has_content(rel_sec),
        },
        "validation_report": {
            "count": val_sec.get("count") or len(val_sec.get("items") or []),
            "approved": val_sec.get("approved"),
            "rejected": val_sec.get("rejected"),
        },
    }


@app.get("/api/patterns/knowledge")
def patterns_knowledge(timeframe: str | None = None) -> dict[str, Any]:
    symbol, primary_tf = _primary()
    tf = timeframe or None
    kb = PatternKnowledgeBase()
    return kb.summary(symbol, tf)


@app.post("/api/patterns/relations/rebuild")
def patterns_relations_rebuild(req: RunRequest) -> dict[str, Any]:
    """Rebuild pattern relation graph from features for one or more timeframes."""
    symbol, primary_tf = _primary()
    tfs = [str(t).upper() for t in (req.timeframes or []) if t]
    if not tfs:
        tfs = [primary_tf]
    results: dict[str, Any] = {}
    for tf in tfs:
        graph = rebuild_relations_from_features(symbol, tf, lookback=None)
        results[tf] = {
            "edges": len(graph.get("edges") or []),
            "nodes": len(graph.get("nodes") or []),
            "sequences": len(graph.get("sequences") or []),
            "summary": graph.get("summary"),
            "empty": not relations_has_content(graph),
        }
    return {"symbol": symbol, "timeframes": results}


@app.get("/api/patterns/overlay")
def patterns_overlay(timeframe: str | None = None, limit: int = 100) -> dict[str, Any]:
    """Live MT5 overlay snapshot + recent pattern history (Explainable AI)."""
    from atis.shared.mt5_pattern_overlay.bridge import local_overlay_dir
    from atis.shared.mt5_pattern_overlay.history import read_history
    from atis.shared.mt5_pattern_overlay.service import get_overlay_service

    symbol, primary_tf = _primary()
    tf = (timeframe or primary_tf or "H1").upper()
    state_path = local_overlay_dir() / "overlay_state.json"
    named = local_overlay_dir() / f"overlay_{symbol}_{tf}.json"
    payload: dict[str, Any] = {}
    for candidate in (named, state_path):
        if candidate.exists():
            try:
                payload = json.loads(candidate.read_text(encoding="utf-8"))
                break
            except Exception:
                payload = {}
    svc = get_overlay_service(load_engine_config().get("engine5_live", {}))
    active = [p.to_dict() for p in svc.active_overlays()]
    return {
        "symbol": symbol,
        "timeframe": tf,
        "state_file": str(state_path),
        "snapshot": payload,
        "active": active,
        "history": read_history(symbol=symbol, timeframe=tf, limit=limit),
        "legend": payload.get("legend") or [],
    }


@app.get("/api/patterns/files")
def patterns_files(timeframe: str | None = None) -> dict[str, Any]:
    """List saved pattern JSON files by section and timeframe."""
    symbol, _ = _primary()
    files = list_pattern_files(symbol=symbol, timeframe=timeframe)
    return {
        "symbol": symbol,
        "timeframe": timeframe,
        "root": str(patterns_root()),
        "sections": list(SECTIONS),
        "files": files,
        "count": len(files),
    }


@app.get("/api/patterns/files/{timeframe}/{section}")
def patterns_file_section(timeframe: str, section: str) -> dict[str, Any]:
    symbol, _ = _primary()
    if section not in SECTIONS:
        raise HTTPException(400, f"Unknown section. Allowed: {list(SECTIONS)}")
    payload = load_section(symbol, timeframe, section)
    if payload is None:
        raise HTTPException(404, f"No JSON for {symbol}/{timeframe}/{section}")

    from atis.shared.pattern_store import section_path as _section_path

    out = dict(payload)
    items = out.get("items") or []
    truncated = False
    total_count = len(items) if isinstance(items, list) else 0
    if isinstance(items, list) and total_count > 800:
        out["items"] = items[:800]
        truncated = True
        out["total_count"] = total_count
    out["truncated"] = truncated
    out["browseable"] = True
    out["path"] = str(_section_path(symbol, timeframe, section))
    out["relative_path"] = f"data/patterns/{symbol}/{timeframe}/{section}.json"
    return out


@app.post("/api/patterns/export-json")
def patterns_export_json(req: RunRequest) -> dict[str, Any]:
    """Export current KB into per-section / per-timeframe JSON files."""
    symbols, tfs = _resolve_symbols_tfs(req)
    if req.timeframes is None:
        tfs = list(load_timeframes().keys())

    def _fn() -> Any:
        return export_pattern_json_from_kb(symbols, tfs)

    job = jobs.submit("pattern_json_export", _fn)
    return job.to_dict()


@app.post("/api/patterns/discover")
def patterns_discover(req: RunRequest) -> dict[str, Any]:
    """Deep pattern exploration — each call runs in its own background thread."""
    symbols, tfs = _resolve_symbols_tfs(req)
    if req.timeframes is None:
        tfs = list(load_timeframes().keys())

    tf_label = ",".join(tfs) if len(tfs) <= 3 else f"{len(tfs)}TFs"
    resume = bool(getattr(req, "resume", True))

    def _fn(job: Any) -> Any:
        def progress(pct: float, message: str) -> None:
            jobs.set_progress(job.id, pct, message)

        def cancel_check() -> None:
            jobs.raise_if_cancelled(job.id)

        def details(payload: dict[str, Any]) -> None:
            jobs.update_details(job.id, {"discovery": payload})

        return run_pattern_discovery(
            symbols,
            tfs,
            force_rebuild=bool(req.force_rebuild),
            progress=progress,
            cancel_check=cancel_check,
            details=details,
            resume=resume,
        )

    job = jobs.submit(f"pattern_discovery:{tf_label}", _fn)
    payload = job.to_dict()
    payload["timeframes"] = tfs
    payload["thread"] = True
    payload["resume"] = resume
    return payload


@app.get("/api/training/details")
def training_details(timeframe: str | None = None) -> dict[str, Any]:
    """Training dataset + metrics for one TF, plus summary for all gold timeframes."""
    symbol, primary_tf = _primary()
    tfs = list(
        load_engine_config().get("engine4_training", {}).get("default_timeframes")
        or load_timeframes().keys()
    )

    def _details_for(tframe: str) -> dict[str, Any]:
        source_meta = training_source_meta(symbol, tframe)
        feat_layer = ((source_meta.get("registry") or {}).get("features") or {})
        feat_meta = {
            "exists": source_meta.get("features_json_exists"),
            "rows": source_meta.get("features_json_rows"),
            "path": source_meta.get("features_json_path"),
            "first_ts": feat_layer.get("first_available_ts"),
            "last_ts": feat_layer.get("last_updated_ts"),
            "days": None,
        }
        if feat_meta["first_ts"] and feat_meta["last_ts"]:
            try:
                t0 = pd.to_datetime(feat_meta["first_ts"], utc=True)
                t1 = pd.to_datetime(feat_meta["last_ts"], utc=True)
                feat_meta["days"] = float((t1 - t0).total_seconds() / 86400.0)
            except Exception:
                pass
        model_dir = _latest_model_dir(symbol, tframe)
        if model_dir is None:
            return {
                "symbol": symbol,
                "timeframe": tframe,
                "dataset": {**feat_meta, "data_sources": source_meta},
                "model": None,
                "empty": True,
                "status": "no_model",
            }
        metrics = _read_json(model_dir / "metrics_report.json")
        backtest = _read_json(model_dir / "backtest_report.json")
        features = _read_json(model_dir / "feature_list.json") or []
        meta = _read_json(model_dir / "metadata.json")
        train_cfg = _read_json(model_dir / "training_config.yaml")
        dataset = {
            **(feat_meta or {}),
            "n_rows_used": (metrics or {}).get("n_rows"),
            "n_features": (metrics or {}).get("n_features") or len(features),
            "feature_list": features,
            "labeling": (train_cfg or {}).get("labeling"),
            "horizon_bars": (metrics or {}).get("horizon_bars") or (train_cfg or {}).get("horizon_bars"),
            "model_type": (metrics or {}).get("model") or (train_cfg or {}).get("baseline_model"),
            "data_sources": source_meta,
        }
        return {
            "symbol": symbol,
            "timeframe": tframe,
            "version": model_dir.name,
            "artifact_dir": str(model_dir),
            "artifact_path": str(model_dir / "model.joblib"),
            "dataset": dataset,
            "metadata": meta,
            "metrics": metrics,
            "backtest": backtest,
            "training_config": train_cfg,
            # Always prefer metrics_report for the latest artifact; never OR with stale meta True.
            "passed_gates": bool((metrics or {}).get("passed_gates", False)),
            "gate_failures": (metrics or {}).get("gate_failures") or [],
            "empty": False,
            "status": "ready",
        }

    all_tf = {t: _details_for(t) for t in tfs}

    def _pick_default_tf() -> str:
        if timeframe:
            return timeframe
        final_tf = str((_final_model_meta() or {}).get("timeframe") or "")
        ready = [t for t in tfs if not (all_tf.get(t) or {}).get("empty", True)]
        if not ready:
            return primary_tf if primary_tf in tfs else (tfs[0] if tfs else primary_tf)
        gated = [t for t in ready if (all_tf.get(t) or {}).get("passed_gates")]
        if final_tf in ready:
            return final_tf
        if gated:
            def gscore(t: str) -> float:
                fin = ((all_tf[t].get("metrics") or {}).get("financial_oos") or {})
                return float(fin.get("sharpe") if fin.get("sharpe") is not None else -1e9)
            return max(gated, key=gscore)

        def score(t: str) -> float:
            fin = ((all_tf[t].get("metrics") or {}).get("financial_oos") or {})
            sharpe = fin.get("sharpe")
            return float(sharpe) if sharpe is not None else -1e9

        return max(ready, key=score)

    tf = _pick_default_tf()
    selected = all_tf.get(tf) or _details_for(tf)
    llmodel = _llmodel_meta()

    # Compact matrix for UI table — champion/latest artifacts (NOT current-run only).
    matrix = []
    for t in tfs:
        item = all_tf[t]
        fin = ((item.get("metrics") or {}).get("financial_oos") or {})
        val = ((item.get("metrics") or {}).get("financial_validation") or {})
        dep = ((item.get("metrics") or {}).get("financial_deploy_holdout") or {})
        cls = ((item.get("metrics") or {}).get("classification") or {})
        diag = ((item.get("metrics") or {}).get("fit_diagnosis") or {})
        ds = item.get("dataset") or {}
        gates = (item.get("metrics") or {}).get("gate_failures") or item.get("gate_failures") or []
        gates_detail = (item.get("metrics") or {}).get("gate_failures_detail") or []
        if not gates_detail and gates:
            try:
                from atis.engines.engine4_training import annotate_gate_failures

                gates_detail = annotate_gate_failures(list(gates))
            except Exception:
                gates_detail = [{"key": g, "ar": str(g), "en": str(g)} for g in gates]
        matrix.append(
            {
                "timeframe": t,
                "status": item.get("status"),
                "empty": item.get("empty", True),
                "version": item.get("version"),
                "pipeline_version": (item.get("metrics") or {}).get("pipeline_version"),
                "passed_gates": item.get("passed_gates", False),
                "gate_failures": gates,
                "gate_failures_detail": gates_detail,
                "rows": ds.get("n_rows_used") or ds.get("rows"),
                "features": ds.get("n_features"),
                "first_ts": ds.get("first_ts"),
                "last_ts": ds.get("last_ts"),
                "days": ds.get("days"),
                "accuracy": cls.get("accuracy"),
                "f1": cls.get("f1_macro"),
                "auc": cls.get("roc_auc_ovr"),
                "trade_rate": cls.get("trade_rate_filtered"),
                "sharpe": fin.get("sharpe"),
                "sharpe_uncapped": fin.get("sharpe_uncapped"),
                "sharpe_ci_low": fin.get("sharpe_ci_low"),
                "ann_factor": fin.get("ann_factor"),
                "mean_trade_return": fin.get("mean_trade_return"),
                "sum_trade_returns": fin.get("sum_trade_returns"),
                "simple_trade_equity": fin.get("simple_trade_equity"),
                "val_sharpe": val.get("sharpe"),
                "deploy_sharpe": dep.get("sharpe"),
                "deploy_trades": dep.get("n_trades"),
                "n_trades_test": fin.get("n_trades"),
                "max_drawdown": fin.get("max_drawdown"),
                "win_rate": fin.get("win_rate"),
                "total_return": fin.get("total_return"),
                "expectancy": fin.get("expectancy"),
                "sortino": fin.get("sortino"),
                "profit_factor": fin.get("profit_factor"),
                "validation_mode": (item.get("metrics") or {}).get("validation_mode")
                or ((item.get("metrics") or {}).get("validation") or {}).get("validation_mode"),
                "regime_validation": _reconcile_regime_validation(
                    (item.get("metrics") or {}).get("regime_validation") or {}
                ),
                "advanced_eval": (item.get("metrics") or {}).get("advanced_eval") or {},
                "knowledge_loop": (item.get("metrics") or {}).get("knowledge_loop") or {},
                "live_readiness": (item.get("metrics") or {}).get("live_readiness") or {},
                "model_zoo": (item.get("metrics") or {}).get("model_zoo") or {},
                "stress_testing": (item.get("metrics") or {}).get("stress_testing") or {},
                "monte_carlo": (item.get("metrics") or {}).get("monte_carlo") or {},
                "intelligent_critique": (item.get("metrics") or {}).get("intelligent_critique") or {},
                "self_optimize": (item.get("metrics") or {}).get("self_optimize") or {},
                "self_optimize_applied": (item.get("metrics") or {}).get("self_optimize_applied") or {},
                "label_quality": (item.get("metrics") or {}).get("label_quality") or {},
                "feature_explainability": (item.get("metrics") or {}).get("feature_explainability") or {},
                "champion_challenger": (item.get("metrics") or {}).get("champion_challenger") or {},
                "smart_recommendations": (item.get("metrics") or {}).get("smart_recommendations") or {},
                "nested_hp": (item.get("metrics") or {}).get("nested_hp") or {},
                "fit_status": diag.get("status"),
                "acc_gap": diag.get("accuracy_gap_train_val"),
                "sharpe_gap_tv": diag.get("sharpe_gap_train_val"),
                "sharpe_gap_vt": diag.get("sharpe_gap_val_test"),
                "source": "champion_or_latest",
                "error": None if not item.get("empty") else "no_model",
            }
        )

    last_run = _read_json(PROJECT_ROOT / "logs" / "training" / "training_run_report.json")
    final_model = _final_model_meta()

    # Current-run matrix from last training_run_report only (never mix with old champions).
    current_run_matrix = []
    run_tfs = list((last_run or {}).get("timeframes") or [])
    run_status = (last_run or {}).get("timeframes_status") or {}
    run_by_tf = {
        r.get("timeframe"): r for r in ((last_run or {}).get("results") or []) if r.get("timeframe")
    }
    for t in (run_tfs or tfs):
        st = run_status.get(t) or {}
        r = run_by_tf.get(t)
        if r is None and not st:
            current_run_matrix.append(
                {
                    "timeframe": t,
                    "status": "not_trained_this_run",
                    "empty": True,
                    "passed_gates": None,
                    "source": "current_run",
                    "label_ar": "لم يُدرَّب في هذا التشغيل",
                }
            )
            continue
        metrics = (r or {}).get("metrics") or {}
        fin = metrics.get("financial_oos") or (st.get("metrics") or {})
        cls = metrics.get("classification") or {}
        diag = metrics.get("fit_diagnosis") or st.get("fit_diagnosis") or {}
        gates = metrics.get("gate_failures") or st.get("gate_failures") or []
        gates_detail = metrics.get("gate_failures_detail") or st.get("gate_failures_detail") or []
        if not gates_detail and gates:
            try:
                from atis.engines.engine4_training import annotate_gate_failures

                gates_detail = annotate_gate_failures(list(gates))
            except Exception:
                gates_detail = [{"key": g, "ar": str(g), "en": str(g)} for g in gates]
        mm = st.get("metrics") or {}
        current_run_matrix.append(
            {
                "timeframe": t,
                "status": "error" if (r or {}).get("error") else (
                    "passed" if (r or {}).get("passed_gates") else (
                        "rejected" if r is not None else st.get("stage") or "done"
                    )
                ),
                "empty": False,
                "version": (r or {}).get("version") or st.get("model_version"),
                "pipeline_version": (last_run or {}).get("pipeline_version") or metrics.get("pipeline_version"),
                "passed_gates": (r or {}).get("passed_gates"),
                "gate_failures": gates,
                "gate_failures_detail": gates_detail,
                "accuracy": cls.get("accuracy") or mm.get("acc"),
                "f1": cls.get("f1_macro") or mm.get("f1"),
                "auc": cls.get("roc_auc_ovr") or mm.get("auc"),
                "trade_rate": cls.get("trade_rate_filtered") or mm.get("trade_rate_filtered"),
                "sharpe": fin.get("sharpe") if isinstance(fin, dict) else mm.get("sharpe"),
                "sharpe_uncapped": (fin.get("sharpe_uncapped") if isinstance(fin, dict) else None)
                or mm.get("sharpe_uncapped"),
                "sharpe_ci_low": (fin.get("sharpe_ci_low") if isinstance(fin, dict) else None)
                or mm.get("sharpe_ci_low"),
                "val_sharpe": ((metrics.get("financial_validation") or {}).get("sharpe") or mm.get("val_sharpe")),
                "deploy_sharpe": (
                    (metrics.get("financial_deploy_holdout") or {}).get("sharpe") or mm.get("deploy_sharpe")
                ),
                "deploy_trades": (
                    (metrics.get("financial_deploy_holdout") or {}).get("n_trades") or mm.get("n_trades_deploy")
                ),
                "n_trades_test": (fin.get("n_trades") if isinstance(fin, dict) else None) or mm.get("n_trades_test"),
                "mean_trade_return": (fin.get("mean_trade_return") if isinstance(fin, dict) else None)
                or mm.get("mean_trade_return"),
                "sum_trade_returns": (fin.get("sum_trade_returns") if isinstance(fin, dict) else None)
                or mm.get("sum_trade_returns"),
                "max_drawdown": (fin.get("max_drawdown") if isinstance(fin, dict) else None)
                or mm.get("max_drawdown"),
                "total_return": (fin.get("total_return") if isinstance(fin, dict) else None),
                "expectancy": (fin.get("expectancy") if isinstance(fin, dict) else None) or mm.get("expectancy"),
                "sortino": (fin.get("sortino") if isinstance(fin, dict) else None) or mm.get("sortino"),
                "profit_factor": (fin.get("profit_factor") if isinstance(fin, dict) else None)
                or mm.get("profit_factor"),
                "validation_mode": metrics.get("validation_mode")
                or st.get("validation_mode")
                or (metrics.get("validation") or {}).get("validation_mode"),
                "regime_validation": _reconcile_regime_validation(
                    metrics.get("regime_validation") or st.get("regime_validation") or {}
                ),
                "advanced_eval": metrics.get("advanced_eval") or st.get("advanced_eval") or {},
                "knowledge_loop": metrics.get("knowledge_loop") or st.get("knowledge_loop") or {},
                "live_readiness": metrics.get("live_readiness") or st.get("live_readiness") or {},
                "model_zoo": metrics.get("model_zoo") or st.get("model_zoo") or {},
                "stress_testing": metrics.get("stress_testing") or st.get("stress_testing") or {},
                "monte_carlo": metrics.get("monte_carlo") or st.get("monte_carlo") or {},
                "intelligent_critique": metrics.get("intelligent_critique")
                or st.get("intelligent_critique")
                or {},
                "self_optimize": metrics.get("self_optimize") or st.get("self_optimize") or {},
                "self_optimize_applied": metrics.get("self_optimize_applied")
                or st.get("self_optimize_applied")
                or {},
                "label_quality": metrics.get("label_quality") or st.get("label_quality") or {},
                "feature_explainability": metrics.get("feature_explainability")
                or st.get("feature_explainability")
                or {},
                "champion_challenger": metrics.get("champion_challenger")
                or st.get("champion_challenger")
                or {},
                "smart_recommendations": metrics.get("smart_recommendations")
                or st.get("smart_recommendations")
                or {},
                "nested_hp": metrics.get("nested_hp") or st.get("nested_hp") or {},
                "fit_status": diag.get("status"),
                "acc_gap": diag.get("accuracy_gap_train_val") or mm.get("acc_gap"),
                "sharpe_gap_tv": diag.get("sharpe_gap_train_val") or mm.get("gap_tv"),
                "sharpe_gap_vt": diag.get("sharpe_gap_val_test") or mm.get("gap_vt"),
                "folds": metrics.get("folds") or st.get("folds") or [],
                "htf_sources": st.get("htf_sources") or [],
                "n_htf_cols": st.get("n_htf_cols"),
                "liquidity_rescue": st.get("liquidity_rescue"),
                "error": (r or {}).get("error"),
                "source": "current_run",
            }
        )

    run_summary = (last_run or {}).get("summary") or {}
    return {
        **selected,
        "selected_timeframe": tf,
        "all_timeframes": tfs,
        "llmodel": llmodel,
        "final_model": final_model,
        "matrix": matrix,
        "matrix_champion": matrix,
        "matrix_current_run": current_run_matrix,
        "last_run": last_run,
        "summary": {
            "total": len(matrix),
            "ready": sum(1 for m in matrix if not m["empty"]),
            "passed_gates": sum(1 for m in matrix if m.get("passed_gates")),
            "missing": sum(1 for m in matrix if m["empty"]),
            "final_model_tf": (final_model or {}).get("timeframe"),
            "final_model_mode": (final_model or {}).get("mode"),
            "final_model_from_prior_run": bool(
                (final_model or {}).get("kept_existing")
                or (final_model or {}).get("skipped_downgrade")
                or (final_model or {}).get("champion_from_prior_run")
            ),
            "current_run_passed_gates": run_summary.get("passed_gates") or run_summary.get("passed"),
            "current_run_trained": run_summary.get("trained"),
            "current_run_rejected": run_summary.get("rejected"),
            "current_run_errors": run_summary.get("errors"),
            "current_run_reject_reasons": run_summary.get("reject_reasons") or [],
            "champion_from_this_run": run_summary.get("champion_from_this_run"),
            "pipeline_version": (last_run or {}).get("pipeline_version"),
            "run_id": (last_run or {}).get("run_id"),
        },
    }


@app.get("/api/overview")
def overview() -> dict[str, Any]:
    symbol, tf = _primary()
    reg = DataStateRegistry()
    layers = {}
    for layer in ("raw", "clean", "features"):
        row = reg.get(symbol, tf, layer)
        layers[layer] = row.__dict__ if row else None

    mt5 = _mt5_status(auto_reconnect=True)

    model_dir = _latest_model_dir(symbol, tf)
    latest_model = None
    if model_dir:
        latest_model = _read_json(model_dir / "metadata.json") or {"version": model_dir.name}

    reports = {
        "ingestion": _read_json(PROJECT_ROOT / "logs" / "ingestion" / "ingestion_run_report.json"),
        "cleaning": _read_json(PROJECT_ROOT / "logs" / "cleaning" / "cleaning_run_report.json"),
        "features": _read_json(PROJECT_ROOT / "logs" / "features" / "features_run_report.json"),
        "training": _read_json(PROJECT_ROOT / "logs" / "training" / "training_run_report.json"),
        "live": _read_json(PROJECT_ROOT / "logs" / "live" / "live_run_report.json"),
        "quality": _read_json(PROJECT_ROOT / "logs" / "cleaning" / "data_quality_report.json"),
    }

    kill = _read_json(PROJECT_ROOT / "logs" / "live" / "kill_switch.json") or {
        "active": bool(load_engine_config().get("engine5_live", {}).get("kill_switch", False))
    }

    price = None
    feat_path = get_path("data_features") / symbol / tf / "features.parquet"
    if feat_path.exists():
        try:
            df = pd.read_parquet(feat_path, columns=["timestamp", "close", "atr"])
            last = df.iloc[-1]
            price = {
                "timestamp": str(last["timestamp"]),
                "close": float(last["close"]),
                "atr": float(last["atr"]) if pd.notna(last["atr"]) else None,
            }
        except Exception:
            price = None

    return {
        "instrument": "Gold",
        "symbol": symbol,
        "timeframe": tf,
        "mt5": mt5,
        "layers": layers,
        "model": latest_model,
        "reports": reports,
        "kill_switch": kill,
        "price": price,
        "jobs": [j.to_dict() for j in jobs.list(10)],
        "autotrader": autotrader.status(),
        "live_settings": _live_settings_payload(),
    }


@app.get("/api/reports/{engine}")
def report(engine: str) -> Any:
    mapping = {
        "ingestion": PROJECT_ROOT / "logs" / "ingestion" / "ingestion_run_report.json",
        "cleaning": PROJECT_ROOT / "logs" / "cleaning" / "cleaning_run_report.json",
        "quality": PROJECT_ROOT / "logs" / "cleaning" / "data_quality_report.json",
        "features": PROJECT_ROOT / "logs" / "features" / "features_run_report.json",
        "training": PROJECT_ROOT / "logs" / "training" / "training_run_report.json",
        "live": PROJECT_ROOT / "logs" / "live" / "live_run_report.json",
        "leaderboard": get_path("models") / "leaderboard.json",
    }
    if engine not in mapping:
        raise HTTPException(404, f"Unknown report: {engine}")
    data = _read_json(mapping[engine])
    if data is None:
        return {"empty": True}
    return data


@app.get("/api/models")
def models() -> dict[str, Any]:
    symbol, tf = _primary()
    base = get_path("models") / symbol / tf
    items = []
    if base.exists():
        for p in sorted([x for x in base.iterdir() if x.is_dir()], reverse=True):
            meta = _read_json(p / "metadata.json") or {"version": p.name}
            metrics = _read_json(p / "metrics_report.json")
            backtest = _read_json(p / "backtest_report.json")
            features = _read_json(p / "feature_list.json")
            items.append(
                {
                    "path": str(p),
                    "meta": meta,
                    "metrics": metrics,
                    "backtest": backtest,
                    "features": features,
                }
            )
    return {
        "symbol": symbol,
        "timeframe": tf,
        "llmodel": _llmodel_meta(),
        "final_model": _final_model_meta(),
        "champion": _read_json(base / "champion.json"),
        "versions": items,
        "leaderboard": _read_json(get_path("models") / "leaderboard.json") or [],
    }


@app.get("/api/llmodel")
def llmodel_status() -> dict[str, Any]:
    symbol, tf = _primary()
    llmodel = _llmodel_meta()
    if llmodel is None:
        return {"exists": False, "symbol": symbol, "timeframe": tf}
    return {
        "exists": True,
        "symbol": symbol,
        "timeframe": tf,
        **llmodel,
    }


@app.get("/api/trades")
def trades(limit: int = 100) -> dict[str, Any]:
    path = PROJECT_ROOT / "logs" / "live" / "trades_log.jsonl"
    rows: list[Any] = []
    if path.exists():
        lines = path.read_text(encoding="utf-8").strip().splitlines()
        for line in lines[-limit:]:
            try:
                rows.append(json.loads(line))
            except json.JSONDecodeError:
                continue
    rows.reverse()
    return {"trades": rows, "count": len(rows)}


@app.get("/api/trades/winning")
def winning_trades(limit: int = 100, symbol: str | None = None, timeframe: str | None = None) -> dict[str, Any]:
    """Winning closed trades with full pattern details — training corpus."""
    try:
        from atis.shared.winning_trade_store import (
            load_winning_trades,
            open_trades_path,
            winning_trades_path,
        )

        sym = symbol or _primary()[0]
        rows = load_winning_trades(symbol=sym, timeframe=timeframe, limit=limit)
        rows = list(reversed(rows))
        return {
            "trades": rows,
            "count": len(rows),
            "symbol": sym,
            "timeframe": timeframe,
            "path": str(winning_trades_path()),
            "open_registry": str(open_trades_path()),
        }
    except Exception as exc:
        raise HTTPException(500, f"تعذر قراءة الصفقات الرابحة: {exc}") from exc


@app.get("/api/rl/monitor")
def rl_monitor(episode_limit: int = 40, timeline_limit: int = 60) -> dict[str, Any]:
    """Learning Monitor: rewards, penalties, knowledge status, training queue."""
    try:
        from atis.shared.rl_learning import get_monitor_snapshot

        return get_monitor_snapshot(
            episode_limit=max(5, min(int(episode_limit), 200)),
            timeline_limit=max(5, min(int(timeline_limit), 300)),
        )
    except Exception as exc:
        raise HTTPException(500, f"تعذر قراءة مراقب التعلم التعزيزي: {exc}") from exc


@app.get("/api/rl/episodes")
def rl_episodes(
    page: int = 1,
    page_size: int = 10,
    limit: int | None = None,
    symbol: str | None = None,
    timeframe: str | None = None,
    kind: str | None = None,
    knowledge_status: str | None = None,
    ticket: str | None = None,
) -> dict[str, Any]:
    """Paginated evaluated RL episodes (reward/penalty). Prefer page/page_size."""
    try:
        from atis.shared.rl_learning import query_episodes

        # Legacy clients that only pass limit still work (single page).
        if limit is not None and page_size == 10 and page == 1:
            page_size = max(1, min(int(limit), 500))
        result = query_episodes(
            page=max(1, int(page)),
            page_size=max(1, min(int(page_size), 100)),
            symbol=symbol,
            timeframe=timeframe,
            kind=kind,
            knowledge_status=knowledge_status,
            ticket=ticket,
        )
        return {**result, "count": int(result.get("total") or 0)}
    except Exception as exc:
        raise HTTPException(500, f"تعذر قراءة حلقات RL: {exc}") from exc


@app.post("/api/rl/episodes/delete")
def rl_episodes_delete(req: DeleteRlEpisodesRequest) -> dict[str, Any]:
    """Delete one or more evaluated RL episodes by episode_id."""
    ids = [str(x).strip() for x in (req.episode_ids or []) if str(x).strip()]
    if not ids:
        raise HTTPException(400, "لم يُمرَّر أي معرف حلقة للحذف")
    try:
        from atis.shared.rl_learning import delete_episodes

        result = delete_episodes(ids)
        return {"ok": True, **result}
    except Exception as exc:
        raise HTTPException(500, f"تعذر حذف حلقات RL: {exc}") from exc


def _rl_display_kind(ep: dict[str, Any]) -> str:
    """Match UI: profit → reward, loss → penalty."""
    try:
        net = float(ep.get("net_profit"))
    except (TypeError, ValueError):
        net = float("nan")
    if net == net:  # finite
        if net > 0:
            return "reward"
        if net < 0:
            return "penalty"
    if ep.get("is_winner"):
        return "reward"
    return str(ep.get("reward_kind") or "neutral")


def _rl_kind_ar(kind: str) -> str:
    if kind == "reward":
        return "رابحة · مكافأة"
    if kind == "penalty":
        return "خاسرة · عقوبة"
    if kind == "neutral":
        return "صافي صفر"
    return kind or "—"


def _rl_knowledge_ar(status: str) -> str:
    if status == "saved":
        return "تم حفظها"
    if status == "pending_review":
        return "قيد المراجعة"
    if status == "rejected":
        return "مرفوضة"
    return status or "—"


@app.get("/api/rl/episodes/export")
def rl_episodes_export(
    limit: int = 2000,
    symbol: str | None = None,
    timeframe: str | None = None,
    kind: str | None = None,
    knowledge_status: str | None = None,
) -> StreamingResponse:
    """Export evaluated RL episodes (reward/penalty) to an Excel workbook."""
    import io
    from datetime import datetime, timezone

    try:
        from atis.shared.rl_learning import load_episodes

        rows = load_episodes(
            limit=max(1, min(int(limit), 10000)),
            symbol=symbol,
            timeframe=timeframe,
            kind=kind,
            knowledge_status=knowledge_status,
        )
        rows = list(reversed(rows))
    except Exception as exc:
        raise HTTPException(500, f"تعذر تصدير حلقات RL: {exc}") from exc

    records: list[dict[str, Any]] = []
    for e in rows:
        disp_kind = _rl_display_kind(e)
        try:
            reward = float(e.get("reward_total"))
            reward_pct = round(reward * 100.0, 2) if reward == reward else None
        except (TypeError, ValueError):
            reward_pct = None
        reasons = e.get("reward_reasons") or []
        lessons = e.get("lessons") or []
        if not isinstance(reasons, list):
            reasons = [str(reasons)]
        if not isinstance(lessons, list):
            lessons = [str(lessons)]
        records.append(
            {
                "الوقت": e.get("evaluated_at") or "",
                "التذكرة": e.get("ticket") if e.get("ticket") is not None else "",
                "الرمز": e.get("symbol") or "",
                "الإطار": e.get("timeframe") or "",
                "الاتجاه": e.get("side") or "",
                "النتيجة (صافي)": e.get("net_profit"),
                "النوع": _rl_kind_ar(disp_kind),
                "القيمة %": reward_pct,
                "السبب": " · ".join(str(x) for x in reasons if x),
                "التأثير على النموذج": e.get("impact_hint") or "",
                "الدروس": " · ".join(str(x) for x in lessons if x),
                "حالة المعرفة": _rl_knowledge_ar(str(e.get("knowledge_status") or "")),
                "للتدريب": "نعم" if e.get("added_to_training_kb") else "لا",
                "سبب الإغلاق": e.get("close_reason") or "",
            }
        )

    df = pd.DataFrame(records)
    buf = io.BytesIO()
    try:
        from openpyxl.styles import Alignment
        from openpyxl.utils import get_column_letter

        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="الصفقات المقيّمة")
            ws = writer.sheets["الصفقات المقيّمة"]
            # Widen text columns so full content is readable in Excel.
            widths = {
                "الوقت": 22,
                "التذكرة": 12,
                "الرمز": 10,
                "الإطار": 8,
                "الاتجاه": 8,
                "النتيجة (صافي)": 14,
                "النوع": 16,
                "القيمة %": 10,
                "السبب": 55,
                "التأثير على النموذج": 40,
                "الدروس": 45,
                "حالة المعرفة": 14,
                "للتدريب": 10,
                "سبب الإغلاق": 18,
            }
            wrap = Alignment(wrap_text=True, vertical="top")
            for idx, col in enumerate(df.columns, start=1):
                letter = get_column_letter(idx)
                ws.column_dimensions[letter].width = float(widths.get(str(col), 14))
                for cell in ws[letter]:
                    cell.alignment = wrap
    except Exception as exc:
        raise HTTPException(500, f"تعذر إنشاء ملف Excel: {exc}") from exc

    buf.seek(0)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"rl_evaluated_trades_{stamp}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/rl/training-queue")
def rl_training_queue(limit: int = 100) -> dict[str, Any]:
    try:
        from atis.shared.rl_learning import episodes_pending_for_training

        rows = episodes_pending_for_training(limit=max(1, min(int(limit), 500)))
        return {
            "pending": list(reversed(rows)),
            "count": len(rows),
            "note": "خبرات محفوظة ستُحقن كسياق في دورة التدريب التالية (Engine4)",
        }
    except Exception as exc:
        raise HTTPException(500, f"تعذر قراءة طابور تدريب RL: {exc}") from exc


@app.post("/api/rl/repair")
def rl_repair() -> dict[str, Any]:
    """Re-score stored RL episodes so winners are never shown as penalties."""
    try:
        from atis.shared.rl_learning import get_monitor_snapshot, repair_episodes

        result = repair_episodes(force=True)
        snap = get_monitor_snapshot(episode_limit=5, timeline_limit=5)
        return {"ok": True, "repair": result, "counts": snap.get("counts")}
    except Exception as exc:
        raise HTTPException(500, f"تعذر إصلاح حلقات RL: {exc}") from exc


@app.get("/api/positions")
def positions(
    symbol: str | None = None,
    atis_only: bool = True,
    live: bool = False,
) -> dict[str, Any]:
    """Open MT5 positions from the watcher cache (instant). Use live=1 to force poll."""
    sym = symbol or _primary()[0]
    # Non-default filters go direct to MT5; default path uses the watcher thread.
    if symbol is not None or atis_only is False:
        try:
            with mt5_session() as client:
                rows = list_open_positions(client, sym, atis_only=atis_only)
        except Exception as exc:
            raise HTTPException(503, f"تعذر قراءة الصفقات المفتوحة: {exc}") from exc
        winners = sum(1 for p in rows if float(p.get("net_profit", 0) or 0) > 0)
        losers = sum(1 for p in rows if float(p.get("net_profit", 0) or 0) < 0)
        total_pnl = sum(float(p.get("net_profit", 0) or 0) for p in rows)
        return {
            "positions": rows,
            "count": len(rows),
            "winners": winners,
            "losers": losers,
            "total_pnl": total_pnl,
            "symbol": sym,
            "atis_only": atis_only,
            "source": "live",
        }
    if live:
        return position_watcher.refresh_now()
    snap = position_watcher.snapshot()
    if snap.get("updated_at") is None:
        return position_watcher.refresh_now()
    if snap.get("error") and not snap.get("positions"):
        raise HTTPException(503, f"تعذر قراءة الصفقات المفتوحة: {snap['error']}")
    return snap


@app.get("/api/positions/stream")
async def positions_stream() -> StreamingResponse:
    """SSE push of open-position snapshots from the watcher thread."""

    async def _events():
        last_seq = -1
        while True:
            snap = position_watcher.snapshot()
            seq = int(snap.get("seq", 0) or 0)
            if seq != last_seq:
                last_seq = seq
                yield f"data: {json.dumps(snap, ensure_ascii=False, default=str)}\n\n"
            await asyncio.sleep(0.15)

    return StreamingResponse(
        _events(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


@app.get("/api/positions/auto-close")
def positions_auto_close_get() -> dict[str, Any]:
    """Current auto-close settings for open positions."""
    return {"ok": True, **position_watcher.auto_close_settings()}


@app.post("/api/positions/auto-close")
def positions_auto_close_set(req: AutoCloseSettingsRequest) -> dict[str, Any]:
    """Enable/disable auto-close and optionally set profit/loss thresholds."""
    if (
        req.enabled is None
        and req.min_profit is None
        and req.loss_enabled is None
        and req.max_loss is None
    ):
        raise HTTPException(400, "حدّد enabled و/أو min_profit و/أو loss_enabled و/أو max_loss")
    try:
        settings = position_watcher.set_auto_close(
            enabled=req.enabled,
            min_profit=req.min_profit,
            loss_enabled=req.loss_enabled,
            max_loss=req.max_loss,
        )
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    return {"ok": True, **settings}


@app.post("/api/positions/close")
def positions_close(req: ClosePositionsRequest) -> dict[str, Any]:
    """Close one open position by ticket, or bulk close by mode."""
    sym = req.symbol or _primary()[0]
    mode = (req.mode or "").strip().lower() or None
    if req.ticket is None and mode is None:
        raise HTTPException(400, "حدّد ticket أو mode (all|winners|losers)")
    if req.ticket is not None and mode is not None:
        raise HTTPException(400, "استخدم إما ticket أو mode وليس الاثنين معاً")
    try:
        with mt5_session() as client:
            if req.ticket is not None:
                result = close_position(client, int(req.ticket))
                if not result.get("ok"):
                    raise HTTPException(
                        400,
                        f"فشل إغلاق الصفقة #{req.ticket}: retcode={result.get('retcode')} {result.get('comment')}",
                    )
                position_watcher.refresh_now()
                return {"ok": True, "closed_count": 1, "closed": [result], "failed": [], "mode": "ticket"}
            assert mode is not None
            if mode not in {"all", "winners", "losers"}:
                raise HTTPException(400, "mode يجب أن يكون all أو winners أو losers")
            result = close_positions_filtered(
                client,
                mode=mode,
                symbol=sym,
                atis_only=bool(req.atis_only),
            )
            if result["matched"] and result["failed_count"] and not result["closed_count"]:
                raise HTTPException(400, f"فشل إغلاق الصفقات ({mode})")
            position_watcher.refresh_now()
            return result
    except HTTPException:
        raise
    except ValueError as exc:
        raise HTTPException(404, str(exc)) from exc
    except Exception as exc:
        raise HTTPException(503, f"تعذر إغلاق الصفقات: {exc}") from exc


@app.get("/api/decision/latest")
def latest_decision() -> dict[str, Any]:
    path = PROJECT_ROOT / "logs" / "live" / "decisions_log.jsonl"
    rows = _read_jsonl_tail(path, 1)
    if not rows:
        return {"empty": True}
    return {"empty": False, "decision": rows[-1]}


@app.get("/api/series/ohlc")
def ohlc(limit: int = 200, layer: str = "features") -> dict[str, Any]:
    symbol, tf = _primary()
    if layer == "raw":
        path = get_path("data_raw") / symbol / tf / f"{symbol}_{tf}.parquet"
    elif layer == "clean":
        path = get_path("data_clean") / symbol / tf / f"{symbol}_{tf}.parquet"
    else:
        path = get_path("data_features") / symbol / tf / "features.parquet"
    if not path.exists():
        raise HTTPException(404, f"No {layer} data for {symbol}/{tf}")
    df = pd.read_parquet(path)
    df = df.tail(limit)
    cols = [c for c in ["timestamp", "open", "high", "low", "close", "tick_volume", "atr", "rsi_14"] if c in df.columns]
    records = df[cols].copy()
    records["timestamp"] = records["timestamp"].astype(str)
    return {"symbol": symbol, "timeframe": tf, "layer": layer, "rows": records.to_dict(orient="records")}


@app.get("/api/jobs")
def list_jobs() -> dict[str, Any]:
    rows = []
    for j in jobs.list():
        d = j.to_dict()
        if isinstance(d.get("details"), dict):
            d["details"] = _reconcile_job_details(d.get("details"))
        rows.append(d)
    return {"jobs": rows}


@app.get("/api/jobs/{job_id}")
def get_job(job_id: str) -> dict[str, Any]:
    job = jobs.get(job_id)
    if not job:
        raise HTTPException(404, "Job not found")
    d = job.to_dict()
    if isinstance(d.get("details"), dict):
        d["details"] = _reconcile_job_details(d.get("details"))
    return d


@app.post("/api/jobs/{job_id}/cancel")
def cancel_job(job_id: str) -> dict[str, Any]:
    job = jobs.cancel(job_id)
    if not job:
        raise HTTPException(404, "Job not found")
    return job.to_dict()


@app.get("/api/autotrade/status")
def autotrade_status() -> dict[str, Any]:
    return autotrader.status()


@app.post("/api/autotrade/start")
def autotrade_start(req: AutoTradeRequest) -> dict[str, Any]:
    symbol, tf = _primary()
    try:
        # Fast path: clear kill switch only when needed (avoid slow full YAML rewrite).
        cfg_live = load_engine_config().get("engine5_live", {}) or {}
        if cfg_live.get("kill_switch"):
            cfg_path = CONFIG_DIR / "engine_config.yaml"
            raw = yaml.safe_load(cfg_path.read_text(encoding="utf-8"))
            raw.setdefault("engine5_live", {})["kill_switch"] = False
            cfg_path.write_text(yaml.safe_dump(raw, sort_keys=False, allow_unicode=True), encoding="utf-8")
            load_engine_config.cache_clear()
        flag = PROJECT_ROOT / "logs" / "live" / "kill_switch.json"
        flag.parent.mkdir(parents=True, exist_ok=True)
        try:
            prev = json.loads(flag.read_text(encoding="utf-8")) if flag.exists() else {}
        except Exception:
            prev = {}
        if prev.get("active"):
            flag.write_text(
                json.dumps(
                    {
                        "active": False,
                        "reason": "autotrade_start",
                        "ts": __import__("datetime").datetime.now(__import__("datetime").timezone.utc).isoformat(),
                    },
                    indent=2,
                ),
                encoding="utf-8",
            )

        tfs = req.timeframes
        if not tfs and req.timeframe:
            tfs = [req.timeframe]
        if not tfs:
            tfs = [tf]
        fusion_mode = req.fusion_mode
        independent = req.multi_tf_independent
        if fusion_mode is None and independent is None and len(tfs) > 1:
            # UI multi-select ⇒ each TF trades with its own model (not fused).
            independent = True
            fusion_mode = "independent"
        return autotrader.start(
            mode=req.mode,
            interval_seconds=req.interval_seconds,
            symbol=req.symbol or symbol,
            timeframe=tfs[0],
            timeframes=tfs,
            multi_tf_independent=independent,
            fusion_mode=fusion_mode,
        )
    except Exception as exc:
        raise HTTPException(400, str(exc)) from exc


@app.post("/api/autotrade/stop")
def autotrade_stop() -> dict[str, Any]:
    return autotrader.stop()


def _resolve_symbols_tfs(req: RunRequest) -> tuple[list[str], list[str]]:
    symbol, tf = _primary()
    symbols = req.symbols or [symbol]
    timeframes = req.timeframes or [tf]
    return symbols, timeframes


@app.post("/api/engines/1/run")
def run_engine1(req: RunRequest) -> dict[str, Any]:
    symbols, tfs = _resolve_symbols_tfs(req)

    def _fn() -> Any:
        # If no explicit TFs beyond default single TF, run all gold TFs when requested via empty override
        return run_ingestion(symbols, tfs, force_rebuild=req.force_rebuild).to_dict()

    job = jobs.submit("engine1_ingestion", _fn)
    return job.to_dict()


@app.post("/api/engines/2/run")
def run_engine2(req: RunRequest) -> dict[str, Any]:
    symbols, tfs = _resolve_symbols_tfs(req)

    def _fn() -> Any:
        return run_cleaning(symbols, tfs, force_rebuild=req.force_rebuild).to_dict()

    job = jobs.submit("engine2_cleaning", _fn)
    return job.to_dict()


@app.post("/api/engines/3/run")
def run_engine3(req: RunRequest) -> dict[str, Any]:
    symbols, tfs = _resolve_symbols_tfs(req)

    def _fn() -> Any:
        return run_features(symbols, tfs, force_rebuild=req.force_rebuild).to_dict()

    job = jobs.submit("engine3_features", _fn)
    return job.to_dict()


@app.post("/api/engines/4/run")
def run_engine4(req: RunRequest) -> dict[str, Any]:
    import importlib
    import sys
    import traceback
    from pathlib import Path

    from atis.config import clear_config_caches
    from atis.shared.logging_utils import get_logger

    log = get_logger("atis.web.engine4")
    clear_config_caches()

    # Reload by fully-qualified name. Do NOT `import ...multi_tf_decision as m`:
    # package __init__ re-exports a function with the same name, which shadows the
    # submodule and makes importlib.reload() raise ImportError → HTTP 500.
    # Include new TIE modules before the package so reload order stays acyclic.
    _E4_RELOAD = (
        "atis.engines.engine4_training.validation_protocols",
        "atis.engines.engine4_training.advanced_metrics",
        "atis.engines.engine4_training.execution_realism",
        "atis.engines.engine4_training.ensemble_models",
        "atis.engines.engine4_training.knowledge_loop",
        "atis.engines.engine4_training.data_quality_gate",
        "atis.engines.engine4_training.adaptive_learning",
        "atis.engines.engine4_training.intelligence",
        "atis.engines.engine4_training.data_intelligence",
        "atis.engines.engine4_training.feature_intelligence",
        "atis.engines.engine4_training.label_quality",
        "atis.engines.engine4_training.feature_explainability",
        "atis.engines.engine4_training.champion_challenger",
        "atis.engines.engine4_training.smart_recommendations",
        "atis.engines.engine4_training.financial_hpo",
        "atis.engines.engine4_training.barrier_optimization",
        "atis.engines.engine4_training.promotion_v16",
        "atis.engines.engine4_training.research_factory",
        "atis.engines.engine4_training.shadow_challenger",
        "atis.engines.engine4_training.enterprise_report",
        "atis.engines.engine4_training.readiness",
        "atis.engines.engine4_training.model_zoo",
        "atis.engines.engine4_training.stress_testing",
        "atis.engines.engine4_training.data_sources",
        "atis.engines.engine4_training.final_model",
        "atis.engines.engine4_training.multi_tf_decision",
        "atis.engines.engine4_training.deep_learning",
        "atis.engines.engine4_training",
    )

    def _reload_engine4() -> Any:
        clear_config_caches()
        last_err: Exception | None = None
        for name in _E4_RELOAD:
            try:
                importlib.import_module(name)
                mod = sys.modules.get(name)
                if mod is not None and getattr(mod, "__spec__", None) is not None:
                    # Avoid reloading a submodule that was shadowed by a same-named
                    # attribute on the parent package (ImportError: not a package module).
                    if name.rsplit(".", 1)[-1] in getattr(
                        sys.modules.get("atis.engines.engine4_training"), "__dict__", {}
                    ) and not name.endswith("engine4_training"):
                        parent = sys.modules.get("atis.engines.engine4_training")
                        attr = getattr(parent, name.rsplit(".", 1)[-1], None)
                        if attr is not None and not hasattr(attr, "__spec__"):
                            # Parent holds a function/object, not the module — reload via sys.modules only.
                            pass
                    importlib.reload(mod)
            except Exception as exc:
                last_err = exc
                if name.endswith(".deep_learning"):
                    continue
                # Soft-fail optional TIE helpers; hard-fail core package.
                if any(
                    name.endswith(sfx)
                    for sfx in (
                        ".validation_protocols",
                        ".advanced_metrics",
                        ".execution_realism",
                        ".ensemble_models",
                        ".knowledge_loop",
                        ".intelligence",
                        ".adaptive_learning",
                        ".data_quality_gate",
                        ".label_quality",
                        ".feature_explainability",
                        ".champion_challenger",
                        ".smart_recommendations",
                        ".financial_hpo",
                        ".barrier_optimization",
                        ".promotion_v16",
                        ".research_factory",
                        ".shadow_challenger",
                        ".enterprise_report",
                        ".data_intelligence",
                        ".feature_intelligence",
                        ".readiness",
                        ".model_zoo",
                        ".stress_testing",
                    )
                ):
                    log.warning("engine4_reload_optional_failed", module=name, error=str(exc))
                    continue
                raise
        # Force-bind critical submodules after package reload (avoid stale function objects).
        for critical in (
            "atis.engines.engine4_training.stress_testing",
            "atis.engines.engine4_training.model_zoo",
            "atis.engines.engine4_training.financial_hpo",
            "atis.engines.engine4_training.barrier_optimization",
        ):
            try:
                importlib.import_module(critical)
                mod = sys.modules.get(critical)
                if mod is not None:
                    importlib.reload(mod)
            except Exception as exc:
                log.warning("engine4_reload_critical_retry_failed", module=critical, error=str(exc))
        if "atis.engines.engine4_training" not in sys.modules:
            raise RuntimeError(f"engine4 package missing after reload: {last_err}")
        return sys.modules["atis.engines.engine4_training"]

    try:
        e4_mod = _reload_engine4()
    except Exception as exc:
        tb = traceback.format_exc()
        try:
            err_path = Path(load_engine_config().get("paths", {}).get("logs", "logs")) / "training" / "engine4_api_error.txt"
            err_path.parent.mkdir(parents=True, exist_ok=True)
            err_path.write_text(tb, encoding="utf-8")
        except Exception:
            pass
        log.exception("engine4_api_reload_failed", error=str(exc))
        raise HTTPException(status_code=500, detail=f"engine4 reload failed: {exc}") from exc

    try:
        symbol, _ = _primary()
        cfg_e4 = load_engine_config().get("engine4_training", {})
        symbols = req.symbols or list(cfg_e4.get("default_symbols") or [symbol])
        # Default: train ALL gold timeframes including M1/M5
        timeframes = req.timeframes or list(
            cfg_e4.get("default_timeframes")
            or ["M1", "M5", "M15", "M30", "H1", "H4"]
        )

        def _fn(job: Any) -> Any:
            def emit_progress(pct: float, message: str) -> None:
                jobs.set_progress(job.id, pct, message)

            def emit_log(line: str) -> None:
                jobs.append_log(job.id, line)

            def emit_status(payload: dict[str, Any]) -> None:
                # Keep structured current-run status on the job for live UI polling.
                event = str((payload or {}).get("event") or "tf_update")
                if event == "run_start":
                    jobs.set_details(
                        job.id,
                        {
                            "pipeline_version": payload.get("pipeline_version"),
                            "run_id": payload.get("run_id"),
                            "timeframes": payload.get("timeframes") or {},
                            "summary": {},
                        },
                    )
                elif event in {"tf_update", "run_done"}:
                    patch: dict[str, Any] = {
                        "pipeline_version": payload.get("pipeline_version"),
                        "run_id": payload.get("run_id"),
                    }
                    if payload.get("timeframes") is not None:
                        patch["timeframes"] = payload.get("timeframes")
                    if payload.get("summary") is not None:
                        patch["summary"] = payload.get("summary")
                    if payload.get("tf") is not None and payload.get("timeframe"):
                        job_obj = jobs.get(job.id)
                        tfs = dict(((job_obj.details if job_obj else None) or {}).get("timeframes") or {})
                        tfs[str(payload.get("timeframe"))] = payload.get("tf")
                        patch["timeframes"] = tfs
                    jobs.update_details(job.id, patch)
                else:
                    jobs.update_details(job.id, payload or {})

            # Reload again inside worker so background jobs see latest code/config.
            mod = _reload_engine4()
            return mod.run_training(
                symbols,
                timeframes,
                progress=emit_progress,
                log=emit_log,
                status=emit_status,
            )

        job = jobs.submit("engine4_training_all_tf", _fn)
        payload = job.to_dict()
        payload["pipeline_version"] = getattr(e4_mod, "PIPELINE_VERSION", None)
        payload["requested_timeframes"] = timeframes
        return payload
    except HTTPException:
        raise
    except Exception as exc:
        tb = traceback.format_exc()
        try:
            err_path = Path("logs") / "training" / "engine4_api_error.txt"
            err_path.parent.mkdir(parents=True, exist_ok=True)
            err_path.write_text(tb, encoding="utf-8")
        except Exception:
            pass
        log.exception("engine4_api_run_failed", error=str(exc))
        raise HTTPException(status_code=500, detail=f"engine4 run failed: {exc}") from exc


@app.post("/api/engines/5/run")
def run_engine5(req: LiveRequest) -> dict[str, Any]:
    symbol, tf = _primary()
    symbols = req.symbols or [symbol]
    allowed = set(load_engine_config().get("trading", {}).get("allowed_live_symbols", ["XAUUSD"]))
    bad = [s for s in symbols if s not in allowed]
    if bad:
        raise HTTPException(400, f"Live trading restricted to gold symbols {sorted(allowed)}; got {bad}")
    timeframe = req.timeframe or tf

    def _fn() -> Any:
        from dataclasses import asdict

        report = run_live_once(
            symbols,
            timeframe,
            dry_run=not req.execute_demo,
            allow_ungated=req.allow_ungated,
        )
        return asdict(report)

    job = jobs.submit("engine5_live", _fn)
    return job.to_dict()


@app.post("/api/pipeline/1-3")
def run_pipeline(req: RunRequest) -> dict[str, Any]:
    symbols, tfs = _resolve_symbols_tfs(req)
    # Default pipeline for gold desk: all timeframes if caller didn't specify
    if req.timeframes is None:
        tfs = list(load_timeframes().keys())

    stages = [
        ("جلب البيانات", run_ingestion),
        ("تنظيف", run_cleaning),
        ("الميزات والأنماط", run_features),
    ]
    total_steps = max(1, len(symbols) * len(tfs) * len(stages))

    def _fn(job: Any) -> Any:
        results: dict[str, Any] = {}
        step = 0
        for stage_name, runner in stages:
            jobs.raise_if_cancelled(job.id)
            stage_key = {
                "جلب البيانات": "engine1",
                "تنظيف": "engine2",
                "الميزات والأنماط": "engine3",
            }[stage_name]
            # Run one TF at a time so the UI can show live %
            stage_reports: list[dict[str, Any]] = []
            for sym in symbols:
                for tf in tfs:
                    jobs.raise_if_cancelled(job.id)
                    jobs.set_progress(
                        job.id,
                        100.0 * step / total_steps,
                        f"{stage_name} · {sym} · {tf} ({step + 1}/{total_steps})",
                    )
                    report = runner([sym], [tf], force_rebuild=req.force_rebuild)
                    jobs.raise_if_cancelled(job.id)
                    stage_reports.append(report.to_dict())
                    step += 1
                    jobs.set_progress(
                        job.id,
                        100.0 * step / total_steps,
                        f"تم: {stage_name} · {sym} · {tf}",
                    )
            results[stage_key] = {
                "items": stage_reports,
                "symbols": symbols,
                "timeframes": tfs,
            }
        jobs.set_progress(job.id, 100.0, "اكتمل تحديث كل الأطر")
        return results

    job = jobs.submit("pipeline_1_3", _fn)
    return job.to_dict()


@app.get("/api/settings/live")
def get_live_settings() -> dict[str, Any]:
    return _live_settings_payload()


@app.post("/api/settings/live")
def update_live_settings(req: LiveSettingsRequest) -> dict[str, Any]:
    cfg_path = CONFIG_DIR / "engine_config.yaml"
    raw = yaml.safe_load(cfg_path.read_text(encoding="utf-8")) or {}
    live = raw.setdefault("engine5_live", {})

    if req.use_live_spread_filter is not None:
        live["use_live_spread_filter"] = bool(req.use_live_spread_filter)
    if req.max_entry_spread_pips is not None:
        live["max_entry_spread_pips"] = float(max(0.1, req.max_entry_spread_pips))
    if req.tight_spread_pips is not None:
        live["tight_spread_pips"] = float(max(0.1, req.tight_spread_pips))
    if req.max_entries_per_cycle is not None:
        live["max_entries_per_cycle"] = int(max(1, min(50, req.max_entries_per_cycle)))

    # Keep tight ≤ max so the gate stays coherent.
    tight = float(live.get("tight_spread_pips", 12.0))
    mx = float(live.get("max_entry_spread_pips", 12.0))
    if tight > mx:
        live["tight_spread_pips"] = mx

    cfg_path.write_text(yaml.safe_dump(raw, sort_keys=False, allow_unicode=True), encoding="utf-8")
    clear_config_caches()
    return _live_settings_payload(live)


@app.get("/api/settings/mt5")
def get_mt5_settings() -> dict[str, Any]:
    payload = _mt5_settings_payload()
    status = _mt5_status(auto_reconnect=False)
    payload["connection"] = status
    return payload


@app.post("/api/settings/mt5")
def update_mt5_settings(req: MT5CredentialsRequest) -> dict[str, Any]:
    current = _mt5_settings_payload()
    login = req.login if req.login is not None and str(req.login).strip() else current.get("login")
    server = req.server if req.server is not None and str(req.server).strip() else current.get("server")
    if not login:
        raise HTTPException(400, "رقم الحساب (Login) مطلوب")
    if not server:
        raise HTTPException(400, "اسم السيرفر مطلوب")

    password = req.password
    keep_pwd = password is None or str(password).strip() == "" or str(password).strip() in {"••••••••", "********"}
    if keep_pwd and not current.get("password_set"):
        raise HTTPException(400, "كلمة المرور مطلوبة")

    try:
        save_mt5_credentials(
            login=login,
            password=None if keep_pwd else str(password),
            server=str(server),
            path=req.path,
            keep_existing_password=keep_pwd,
        )
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    except Exception as exc:
        raise HTTPException(500, f"تعذر حفظ بيانات MT5: {exc}") from exc

    connection: dict[str, Any] | None = None
    connect_error: str | None = None
    if req.reconnect:
        try:
            connection = _reconnect_mt5_keepalive()
        except Exception as exc:
            connect_error = str(exc)
            connection = {"ok": False, "error": connect_error}

    payload = _mt5_settings_payload()
    payload["saved"] = True
    payload["connection"] = connection or _mt5_status(auto_reconnect=False)
    if connect_error:
        payload["connect_error"] = connect_error
    return payload


@app.post("/api/settings/mt5/test")
def test_mt5_settings(req: MT5CredentialsRequest) -> dict[str, Any]:
    """Verify MT5 login using form values (saves first if credentials provided)."""
    current = _mt5_settings_payload()
    login = req.login if req.login is not None and str(req.login).strip() else current.get("login")
    server = req.server if req.server is not None and str(req.server).strip() else current.get("server")
    password = req.password
    keep_pwd = password is None or str(password).strip() == "" or str(password).strip() in {"••••••••", "********"}

    if not login or not server:
        raise HTTPException(400, "رقم الحساب والسيرفر مطلوبان للتحقق")
    if keep_pwd and not current.get("password_set"):
        raise HTTPException(400, "أدخل كلمة المرور أولاً")

    # Persist then reconnect so the shared keepalive matches the verified account.
    try:
        save_mt5_credentials(
            login=login,
            password=None if keep_pwd else str(password),
            server=str(server),
            path=req.path,
            keep_existing_password=keep_pwd,
        )
        connection = _reconnect_mt5_keepalive()
        payload = _mt5_settings_payload()
        payload["saved"] = True
        payload["connection"] = connection
        payload["ok"] = True
        return payload
    except HTTPException:
        raise
    except Exception as exc:
        return {
            **_mt5_settings_payload(),
            "ok": False,
            "saved": True,
            "connection": {"ok": False, "error": str(exc)},
            "error": str(exc),
        }


@app.get("/api/settings/training")
def get_training_settings() -> dict[str, Any]:
    return _training_settings_payload()


@app.post("/api/settings/training")
def update_training_settings(req: TrainingSettingsRequest) -> dict[str, Any]:
    if not isinstance(req.settings, dict) or not req.settings:
        raise HTTPException(400, "settings must be a non-empty object")

    cfg_path = CONFIG_DIR / "engine_config.yaml"
    raw = yaml.safe_load(cfg_path.read_text(encoding="utf-8")) or {}
    # Full replace of engine4_training with the snapshot from the settings UI.
    raw["engine4_training"] = dict(req.settings)
    cfg_path.write_text(yaml.safe_dump(raw, sort_keys=False, allow_unicode=True), encoding="utf-8")
    clear_config_caches()
    return _training_settings_payload(raw["engine4_training"])


@app.post("/api/kill-switch")
def kill_switch(req: KillSwitchRequest) -> dict[str, Any]:
    cfg_path = CONFIG_DIR / "engine_config.yaml"
    raw = yaml.safe_load(cfg_path.read_text(encoding="utf-8"))
    raw.setdefault("engine5_live", {})["kill_switch"] = bool(req.active)
    cfg_path.write_text(yaml.safe_dump(raw, sort_keys=False, allow_unicode=True), encoding="utf-8")
    load_engine_config.cache_clear()

    if req.active:
        autotrader.stop()

    flag = PROJECT_ROOT / "logs" / "live" / "kill_switch.json"
    flag.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "active": bool(req.active),
        "reason": req.reason,
        "ts": __import__("datetime").datetime.now(__import__("datetime").timezone.utc).isoformat(),
    }
    flag.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    return payload
